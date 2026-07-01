"""
SCRIPT  : advisory_revenue_projection_FY2026-27.py
PURPOSE : Generates a professional, formula-driven Excel revenue projection sheet
          for Pruthashakti Kisan Kalyan Mission — Advisory Platform.
CREATED : 2026-06-07
PROJECT : ProfitPortal (PSCMS)

WHAT IT DOES
------------
- Reads quantity data from "Sheet1" in the target Excel workbook.
- Creates a new tab "FY 2026-27 Projection" with:
    * All 3 Advisory Plan tiers (Starter ₹99 / Premium ₹299 / Platinum ₹999)
    * 8 other service lines (Websites, ERP, E-Commerce, KBS, etc.)
    * Advisory Sub-Total and Grand Total rows
    * Cost breakdown section: Dev & IT Maintenance 30%, Server 3%,
      G&A 2%, Procurement 2%, Interest 0.4%
    * Net Revenue row
- Every Revenue, Sub-Total, Total, and Cost cell uses Excel FORMULAS —
  changing "Subscriptions (nos)" or "Rate" rows auto-recalculates everything.

HOW TO REUSE
------------
1. pip install openpyxl          (if not already installed)
2. Update PATH below to point at the new Excel file.
3. Update ADVISORY_PLANS rates if pricing changes.
4. Update OTHER_SERVICES source row numbers if the Sheet1 layout changes.
5. Run: python advisory_revenue_projection_FY2026-27.py

DEPENDENCIES
------------
- Python 3.8+
- openpyxl  (pip install openpyxl)

INPUT  : Excel file at PATH — must have a "Sheet1" with quantity data in rows 9-35.
OUTPUT : Same Excel file with a new "FY 2026-27 Projection" tab added/replaced.
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

PATH = r"D:\PSCMS\ProfitPortal\IT_PROJECTIONS_2026_Professional_Pricing_Model.xlsx"
wb   = openpyxl.load_workbook(PATH)
ws_src = wb["Sheet1"]

Y1 = list(range(4, 16))   # source cols D..O = Apr'26..Mar'27
MONTH_LABELS = [
    "Apr'26","May'26","Jun'26","Jul'26","Aug'26","Sep'26",
    "Oct'26","Nov'26","Dec'26","Jan'27","Feb'27","Mar'27",
]
adv_qty = [ws_src.cell(9, c).value or 0 for c in Y1]

ADVISORY_PLANS = [
    {"plan":"Starter Plan",  "rate":99,  "color":"1E8449",
     "bg_hdr":"D5F5E3","bg_qty":"F0FBF4","bg_rate":"E8F8F0","bg_amt":"D5F5E3"},
    {"plan":"Premium Plan",  "rate":299, "color":"1A5276",
     "bg_hdr":"D6EAF8","bg_qty":"F0F8FF","bg_rate":"E8F4FD","bg_amt":"D6EAF8"},
    {"plan":"Platinum Plan", "rate":999, "color":"5B2C6F",
     "bg_hdr":"E8DAEF","bg_qty":"F9F0FF","bg_rate":"F4E6FD","bg_amt":"E8DAEF"},
]

OTHER_SERVICES = [
    (2,"Websites for FPC / Vendors",          12),
    (3,"Vendor / FPC Registration (ERP)",     15),
    (4,"Pruthashakti Vertical Requirements",  18),
    (5,"E-Commerce Platform (B2B)",           21),
    (6,"E-Commerce Platform (B2C)",           24),
    (7,"KBS Systems",                         27),
    (8,"Global Requirements",                 30),
    (9,"Reference Websites (Agri / Non-Agri)",33),
]

# =============================================================================
SHEET = "FY 2026-27 Projection"
if SHEET in wb.sheetnames:
    del wb[SHEET]
ws = wb.create_sheet(SHEET, 0)

# ── style helpers ─────────────────────────────────────────────────────────────
def fl(h):  return PatternFill("solid", fgColor=h)
def bd(l="thin",r="thin",t="thin",b="thin"):
    def s(v): return Side(style=v) if v else Side(style=None)
    return Border(left=s(l), right=s(r), top=s(t), bottom=s(b))
BORDER  = bd()
BORD_M  = bd("medium","medium","medium","medium")
AC = Alignment(horizontal="center",vertical="center",wrap_text=True)
AL = Alignment(horizontal="left",  vertical="center",indent=1)
AR = Alignment(horizontal="right", vertical="center")
L  = get_column_letter    # shorthand

LAST = 16   # col P = FY Total
DATA = list(range(4, 16))  # cols D..O

def merge_row(r, bg, font_kw=None, text=None, cols=None):
    """Fill a full-width merged label row."""
    c1, c2 = (cols or (1, LAST))
    for c in range(c1, c2+1):
        ws.cell(r, c).fill   = fl(bg)
        ws.cell(r, c).border = BORDER
    ws.merge_cells(f"{L(c1)}{r}:{L(c2)}{r}")
    if text is not None:
        cell = ws.cell(r, c1)
        cell.value     = text
        cell.alignment = AC
        if font_kw:
            cell.font = Font(**{"name":"Calibri","size":10, **font_kw})

def qty_row_write(r, qty_list, color, bg, bg_total):
    ws.row_dimensions[r].height = 14
    ws.cell(r, 3).value     = "  Subscriptions (nos)"
    ws.cell(r, 3).font      = Font(italic=True, color=color, size=9, name="Calibri")
    ws.cell(r, 3).alignment = AL
    for col in range(1, LAST+1):
        ws.cell(r, col).fill   = fl(bg)
        ws.cell(r, col).border = BORDER
    for i, q in enumerate(qty_list):
        c = ws.cell(r, 4+i)
        c.value        = q if q else None
        c.font         = Font(color=color, size=9, name="Calibri")
        c.alignment    = AR
        c.number_format = "#,##0"
    c = ws.cell(r, LAST)
    c.value        = f"=SUM(D{r}:O{r})"
    c.font         = Font(bold=True, color=color, size=9, name="Calibri")
    c.fill         = fl(bg_total)
    c.alignment    = AR
    c.number_format = "#,##0"

def rate_row_write(r, rate_val, color, bg, bg_total):
    ws.row_dimensions[r].height = 14
    ws.cell(r, 3).value     = "  Rate (₹/Acre/Season)" if rate_val else "  Rate (₹/unit)"
    ws.cell(r, 3).font      = Font(italic=True, color=color, size=9, name="Calibri")
    ws.cell(r, 3).alignment = AL
    for col in range(1, LAST+1):
        ws.cell(r, col).fill   = fl(bg)
        ws.cell(r, col).border = BORDER
    for i in range(12):
        c = ws.cell(r, 4+i)
        c.value        = rate_val   # None = TBD
        c.font         = Font(color=color, size=9, name="Calibri")
        c.alignment    = AR
        c.number_format = "₹#,##0"
    c = ws.cell(r, LAST)
    c.value     = "—"
    c.font      = Font(bold=True, color=color, size=9, name="Calibri")
    c.fill      = fl(bg_total)
    c.alignment = AC

def revenue_row_write(r, qty_r, rate_r, color, bg, bg_total, iferror=False):
    ws.row_dimensions[r].height = 14
    ws.cell(r, 3).value     = "  Revenue (₹)"
    ws.cell(r, 3).font      = Font(bold=True, color=color, size=9, name="Calibri")
    ws.cell(r, 3).alignment = AL
    for col in range(1, LAST+1):
        ws.cell(r, col).fill   = fl(bg)
        ws.cell(r, col).border = BORDER
    for i in range(12):
        dc = L(4+i)
        formula = f"={dc}{qty_r}*{dc}{rate_r}"
        if iferror:
            formula = f"=IFERROR({dc}{qty_r}*{dc}{rate_r},0)"
        c = ws.cell(r, 4+i)
        c.value        = formula
        c.font         = Font(bold=True, color=color, size=9, name="Calibri")
        c.alignment    = AR
        c.number_format = "₹#,##0"
    c = ws.cell(r, LAST)
    c.value        = f"=SUM(D{r}:O{r})"
    c.font         = Font(bold=True, color=color, size=9, name="Calibri")
    c.fill         = fl(bg_total)
    c.alignment    = AR
    c.number_format = "₹#,##0"

def grand_row(r, formula_fn, bg, border=BORDER, font_size=10):
    ws.row_dimensions[r].height = 20
    for col in range(1, LAST+1):
        ws.cell(r, col).fill   = fl(bg)
        ws.cell(r, col).border = border
    for i in range(12):
        c = ws.cell(r, 4+i)
        c.value        = formula_fn(L(4+i))
        c.font         = Font(bold=True, color="1A5632", size=font_size, name="Calibri")
        c.alignment    = AR
        c.number_format = "₹#,##0"
    c = ws.cell(r, LAST)
    c.value        = f"=SUM(D{r}:O{r})"
    c.font         = Font(bold=True, color="1A5632", size=font_size+1, name="Calibri")
    c.alignment    = AR
    c.number_format = "₹#,##0"

# ── column widths ─────────────────────────────────────────────────────────────
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 16
for i in range(12):
    ws.column_dimensions[L(4+i)].width = 8.5
ws.column_dimensions[L(LAST)].width = 13

# ──────────────────────────────────────────────────────────────────────────────
#  HEADER BLOCK
# ──────────────────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 34
for c in range(1, LAST+1): ws.cell(1,c).fill = fl("1A5632")
ws.merge_cells(f"A1:{L(LAST)}1")
ws["A1"].value     = "PRUTHASHAKTI KISAN KALYAN MISSION"
ws["A1"].font      = Font(bold=True, color="FFFFFF", size=18, name="Calibri")
ws["A1"].alignment = AC

ws.row_dimensions[2].height = 20
for c in range(1, LAST+1): ws.cell(2,c).fill = fl("27AE60")
ws.merge_cells(f"A2:{L(LAST)}2")
ws["A2"].value     = "Advisory Platform  —  Revenue Projection  |  Financial Year 2026-27  (April 2026 – March 2027)"
ws["A2"].font      = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
ws["A2"].alignment = AC

ws.row_dimensions[3].height = 17
for c in range(1, LAST+1): ws.cell(3,c).fill = fl("D5F5E3")
ws.merge_cells(f"A3:{L(LAST)}3")
ws["A3"].value     = "Advisory Pricing Plans:   STARTER — ₹99/Acre/Season   |   PREMIUM — ₹299/Acre/Season   |   PLATINUM — ₹999/Acre/Season"
ws["A3"].font      = Font(bold=True, color="1A5632", size=9.5, name="Calibri")
ws["A3"].alignment = AC

# Column headers (Row 4)
ws.row_dimensions[4].height = 22
for col_idx, hdr in enumerate(["Sr.","Service / Business Line","Plan / Parameter"]+MONTH_LABELS+["FY 2026-27\nTotal"], 1):
    c = ws.cell(4, col_idx)
    c.value     = hdr
    c.font      = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    c.fill      = fl("2C3E50")
    c.alignment = AC
    c.border    = BORDER

# ──────────────────────────────────────────────────────────────────────────────
#  SECTION 1 – Advisory for Farmers (3 plans)
# ──────────────────────────────────────────────────────────────────────────────
row = 5

# Section title row
ws.row_dimensions[row].height = 18
for c in range(1, LAST+1):
    ws.cell(row,c).fill   = fl("1A5632")
    ws.cell(row,c).border = BORDER
ws.merge_cells(f"B{row}:{L(LAST)}{row}")
ws.cell(row,1).value     = "1"
ws.cell(row,1).font      = Font(bold=True,color="FFFFFF",size=10,name="Calibri")
ws.cell(row,1).alignment = AC
ws.cell(row,2).value     = "Advisory for Farmers"
ws.cell(row,2).font      = Font(bold=True,color="FFFFFF",size=11,name="Calibri")
ws.cell(row,2).alignment = AL
row += 1

plan_amt_rows = []

for p in ADVISORY_PLANS:
    # Plan sub-header
    ws.row_dimensions[row].height = 16
    for c in range(1, LAST+1):
        ws.cell(row,c).fill   = fl(p["bg_hdr"])
        ws.cell(row,c).border = BORDER
    ws.merge_cells(f"B{row}:{L(LAST)}{row}")
    ws.cell(row,2).value     = f"  {p['plan']}  (₹{p['rate']} / Acre / Season)"
    ws.cell(row,2).font      = Font(bold=True, color=p["color"], size=9.5, name="Calibri")
    ws.cell(row,2).alignment = AL
    row += 1; qty_r = row

    qty_row_write(row, adv_qty, p["color"], p["bg_qty"], "FEF9E7")
    row += 1; rate_r = row

    rate_row_write(row, p["rate"], p["color"], p["bg_rate"], "FEF9E7")
    row += 1

    revenue_row_write(row, qty_r, rate_r, p["color"], p["bg_amt"], "FEF9E7")
    plan_amt_rows.append(row)
    row += 1

# Advisory sub-total
adv_sub_row = row
ws.row_dimensions[row].height = 18
for c in range(1, LAST+1):
    ws.cell(row,c).fill   = fl("A9DFBF")
    ws.cell(row,c).border = BORD_M
ws.merge_cells(f"A{row}:C{row}")
ws.cell(row,1).value     = "Advisory Sub-Total Revenue  (Starter + Premium + Platinum)"
ws.cell(row,1).font      = Font(bold=True,color="1A5632",size=9.5,name="Calibri")
ws.cell(row,1).alignment = AL
grand_row(row,
    lambda dc: f"={dc}{plan_amt_rows[0]}+{dc}{plan_amt_rows[1]}+{dc}{plan_amt_rows[2]}",
    "A9DFBF", border=BORD_M)
row += 1

# ──────────────────────────────────────────────────────────────────────────────
#  SECTIONS 2-9 – Other Services
# ──────────────────────────────────────────────────────────────────────────────
other_amt_rows = []

for sr, name, qr_src in OTHER_SERVICES:
    qty = [ws_src.cell(qr_src, c).value or 0 for c in Y1]

    # Service header
    ws.row_dimensions[row].height = 17
    for c in range(1, LAST+1):
        ws.cell(row,c).fill   = fl("EBF5EB")
        ws.cell(row,c).border = BORDER
    ws.merge_cells(f"B{row}:{L(LAST)}{row}")
    ws.cell(row,1).value     = sr
    ws.cell(row,1).font      = Font(bold=True,color="1A5632",size=9,name="Calibri")
    ws.cell(row,1).alignment = AC
    ws.cell(row,2).value     = name
    ws.cell(row,2).font      = Font(bold=True,color="1A5632",size=10,name="Calibri")
    ws.cell(row,2).alignment = AL
    row += 1; qty_r = row

    qty_row_write(row, qty, "1F618D", "FDFEFE", "FEF9E7")
    row += 1; rate_r = row

    rate_row_write(row, None, "6C3483", "F5EEF8", "FEF9E7")
    ws.cell(row, LAST).value = "TBD"
    row += 1

    revenue_row_write(row, qty_r, rate_r, "1A5632", "EBF5FB", "FEF9E7", iferror=True)
    other_amt_rows.append(row)
    row += 1

# ──────────────────────────────────────────────────────────────────────────────
#  GRAND TOTAL – All services
# ──────────────────────────────────────────────────────────────────────────────
row += 1
grand_row_num = row
ws.row_dimensions[row].height = 22
for c in range(1, LAST+1):
    ws.cell(row,c).fill   = fl("D4EFDF")
    ws.cell(row,c).border = BORD_M
ws.merge_cells(f"A{row}:C{row}")
ws.cell(row,1).value     = "TOTAL PROJECTED REVENUE  (All Services)  |  FY 2026-27"
ws.cell(row,1).font      = Font(bold=True,color="1A5632",size=10,name="Calibri")
ws.cell(row,1).alignment = AC

def grand_formula(dc):
    other = "+".join(f"{dc}{r}" for r in other_amt_rows) if other_amt_rows else "0"
    return f"={dc}{adv_sub_row}+{other}"

grand_row(row, grand_formula, "D4EFDF", border=BORD_M, font_size=10)
row += 1

# ──────────────────────────────────────────────────────────────────────────────
#  COST & MARGIN BREAKDOWN
# ──────────────────────────────────────────────────────────────────────────────
row += 1
# Section header
ws.row_dimensions[row].height = 16
for c in range(1, LAST+1):
    ws.cell(row,c).fill   = fl("2C3E50")
    ws.cell(row,c).border = BORDER
ws.merge_cells(f"A{row}:{L(LAST)}{row}")
ws.cell(row,1).value     = "Cost & Margin Breakdown"
ws.cell(row,1).font      = Font(bold=True,color="FFFFFF",size=10,name="Calibri")
ws.cell(row,1).alignment = AC
row += 1

COST_ITEMS = [
    # (label, pct, text_color, row_bg, data_bg)
    ("Development & IT Maintenance", 0.30, "C0392B", "FDEDEC", "FADADD"),
    ("Server & Infrastructure Cost", 0.03, "784212", "FDEBD0", "FAD7A0"),
    ("G&A Expenses",                 0.02, "4A4A4A", "F2F3F4", "EAECEE"),
    ("Procurement Cost",             0.02, "4A4A4A", "F2F3F4", "EAECEE"),
    ("Interest",                     0.004,"4A4A4A", "F2F3F4", "EAECEE"),
]

cost_rows = []
for label, pct, color, bg, data_bg in COST_ITEMS:
    pct_str = f"{pct*100:g}%"
    ws.row_dimensions[row].height = 15
    for c in range(1, LAST+1):
        ws.cell(row,c).fill   = fl(bg)
        ws.cell(row,c).border = BORDER
    ws.merge_cells(f"A{row}:C{row}")
    ws.cell(row,1).value     = f"  {label}  ({pct_str} of Total Revenue)"
    ws.cell(row,1).font      = Font(color=color, size=9, name="Calibri")
    ws.cell(row,1).alignment = AL
    for i in range(12):
        dc = L(4+i)
        c  = ws.cell(row, 4+i)
        c.value        = f"={dc}{grand_row_num}*{pct}"
        c.font         = Font(color=color, size=9, name="Calibri")
        c.fill         = fl(data_bg)
        c.alignment    = AR
        c.number_format = "₹#,##0"
    c = ws.cell(row, LAST)
    c.value        = f"=SUM(D{row}:O{row})"
    c.font         = Font(bold=True, color=color, size=9, name="Calibri")
    c.fill         = fl("FEF9E7")
    c.alignment    = AR
    c.number_format = "₹#,##0"
    cost_rows.append(row)
    row += 1

# Total Costs row
total_cost_row = row
ws.row_dimensions[row].height = 16
for c in range(1, LAST+1):
    ws.cell(row,c).fill   = fl("FADBD8")
    ws.cell(row,c).border = BORD_M
ws.merge_cells(f"A{row}:C{row}")
ws.cell(row,1).value     = "  Total Costs  (30% Dev + 3% Server + 2% G&A + 2% Procurement + 0.4% Interest = 37.4%)"
ws.cell(row,1).font      = Font(bold=True, color="C0392B", size=9, name="Calibri")
ws.cell(row,1).alignment = AL
for i in range(12):
    dc  = L(4+i)
    frm = "+".join(f"{dc}{r}" for r in cost_rows)
    c   = ws.cell(row, 4+i)
    c.value        = f"={frm}"
    c.font         = Font(bold=True, color="C0392B", size=9, name="Calibri")
    c.alignment    = AR
    c.border       = BORD_M
    c.number_format = "₹#,##0"
c = ws.cell(row, LAST)
c.value        = f"=SUM(D{row}:O{row})"
c.font         = Font(bold=True, color="C0392B", size=10, name="Calibri")
c.fill         = fl("FEF9E7")
c.alignment    = AR
c.border       = BORD_M
c.number_format = "₹#,##0"
row += 1

# NET REVENUE row
net_row = row
ws.row_dimensions[row].height = 22
for c in range(1, LAST+1):
    ws.cell(row,c).fill   = fl("D4EFDF")
    ws.cell(row,c).border = BORD_M
ws.merge_cells(f"A{row}:C{row}")
ws.cell(row,1).value     = "NET REVENUE  (Total Revenue − Total Costs)"
ws.cell(row,1).font      = Font(bold=True, color="1A5632", size=10, name="Calibri")
ws.cell(row,1).alignment = AL
for i in range(12):
    dc = L(4+i)
    c  = ws.cell(row, 4+i)
    c.value        = f"={dc}{grand_row_num}-{dc}{total_cost_row}"
    c.font         = Font(bold=True, color="1A5632", size=10, name="Calibri")
    c.alignment    = AR
    c.border       = BORD_M
    c.number_format = "₹#,##0"
c = ws.cell(row, LAST)
c.value        = f"={L(LAST)}{grand_row_num}-{L(LAST)}{total_cost_row}"
c.font         = Font(bold=True, color="1A5632", size=11, name="Calibri")
c.fill         = fl("A9DFBF")
c.alignment    = AR
c.border       = BORD_M
c.number_format = "₹#,##0"
row += 2

# ── FOOTER / INSTRUCTIONS ─────────────────────────────────────────────────────
ws.row_dimensions[row].height = 36
for c in range(1, LAST+1):
    ws.cell(row,c).fill = fl("EBF5EB")
ws.merge_cells(f"A{row}:{L(LAST)}{row}")
ws.cell(row,1).value = (
    "HOW TO USE:  Change any value in the 'Subscriptions (nos)' rows to update the user count — "
    "Revenue, Sub-Totals, Grand Total, and all cost lines recalculate automatically via formulas.  "
    "Change any 'Rate' cell to update pricing instantly.  "
    "Cost breakdown: Dev & IT Maintenance 30%  |  Server & Infrastructure 3%  |  "
    "G&A 2%  |  Procurement 2%  |  Interest 0.4%  =  37.4% total deduction from revenue.  "
    "For other service lines, fill in the Rate rows to unlock their revenue figures."
)
ws.cell(row,1).font      = Font(italic=True, color="1A5632", size=8, name="Calibri")
ws.cell(row,1).alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)

# ── FINAL SETTINGS ────────────────────────────────────────────────────────────
ws.freeze_panes = "D5"
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 90
wb.active = ws
wb.save(PATH)
print("Saved. Formulas embedded — change subscriptions or rate, everything updates.")
print(f"  Sheet '{SHEET}' is the first tab.")
print(f"  Cost breakdown: Dev 30% | Server 3% | G&A 2% | Procurement 2% | Interest 0.4%")
